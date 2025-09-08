import fileinput
import re
import os
from os import path
from pathlib import Path
from openpyxl import load_workbook
from random import randint
import shutil
import sys


def Review_Comment():
    rewarning = []
    result = []
    dataWarning1 = ''
    dataWarning2 = ''
    metricData = ''
    MisraAdded = ''
    flag = 0
    flag2 = 0
    for w_file in os.listdir(inputfolder):    
        WorkingData = open(inputfolder + '\\' + w_file).read()
        #COMMENT WARNING
        #Search dataWarning
        checkwarning = re.search('(totalNoOfWarnings = )+([0-9]+)(;)',WorkingData)
        print('==========================================================================================\n')
        if checkwarning !=None:
            totalNoOfWarnings = checkwarning.group(2)
            if int(totalNoOfWarnings) !=0:            
                print('[PROCESS] Working with file '+ w_file + ': '+str(totalNoOfWarnings)+' warnings!')
                #Search dataWarning
                rewarning = re.search('(dataWarnings=\[\[.+\]\];)', WorkingData)
                result = rewarning.groups(0)
                dataWarning1 = ''.join(result)  
                dataWarning2 = dataWarning1.replace('],','],\n')
                for i in range (2, row_count+1):    
                    if misra[i] == 'none':
                        rewarning = re.search('([\w \(\)\[\]\\~\!\@\#\$\%\^\&*-=+|}{\';:.,?<>\"`\\\\]+)(\"\")([\\\\\w \(\)\[\]\\~\!\@\#\$\%\^\&*-=+|}{\';:.,?<>\"`]+)(' + str(extend[i]) + ')([\\\\\w \(\)\[\]\\~\!\@\#\$\%\^\&*-=+|}{\';:.,?<>\"`]+)(\",\"\",\"\",)(\"[\d]+\"])', dataWarning2)
                        while rewarning != None:
                            flag = 1
                            flag2 = 1
                            result = rewarning.groups(0)
                            concatresult = ''.join(result)
                            dataReplace = str(result[0]) + str(result[1]) + str(result[2]) + str(result[3]) + str(result[4]) +'\",\"'+str(comment[i])+'\",\"' + str(howtosolve[i]) + '\",' + str(result[6])
                            dataWarning2 = dataWarning2.replace(concatresult, dataReplace)
                            rewarning = re.search('([\\\\\w \(\)\[\]\\~\!\@\#\$\%\^\&*-=+|}{\';:.,?<>\"`]+)(\"\")([\\\\\w \(\)\[\]\\~\!\@\#\$\%\^\&*-=+|}{\';:.,?<>\"`]+)(' + str(extend[i]) + ')([\\\\\w \(\)\[\]\\~\!\@\#\$\%\^\&*-=+|}{\';:.,?<>\"`]+)(\",\"\",\"\",)(\"[\d]+\"])', dataWarning2)                           
                    if misra[i] != 'none':
                        rewarning = re.search('(\"'+ str(misra[i]) +'\")([\\\\\w \(\)\[\]\\~\!\@\#\$\%\^\&*-=+|}{\';:.,?<>\"`]+)(' + str(extend[i]) + ')([\\\\\w \(\)\[\]\\~\!\@\#\$\%\^\&*-=+|}{\';:.,?<>\"`]+)(\",\"\",\"\",)(\"[\d]+\"])', dataWarning2)      
                        while rewarning != None:
                            flag = 1
                            flag2 = 1
                            result = rewarning.groups(0)
                            concatresult = ''.join(result)
                            dataReplace = str(result[0]) + str(result[1]) + str(result[2]) + str(result[3]) +'\",\"'+str(comment[i])+'\",\"' + str(howtosolve[i]) + '\",' + str(result[5])
                            dataWarning2 = dataWarning2.replace(concatresult, dataReplace)
                            rewarning = re.search('(\"'+ str(misra[i]) +'\")([\\\\\w \(\)\[\]\\~\!\@\#\$\%\^\&*-=+|}{\';:.,?<>\"`]+)(' + str(extend[i]) + ')([\\\\\w \(\)\[\]\\~\!\@\#\$\%\^\&*-=+|}{\';:.,?<>\"`]+)(\",\"\",\"\",)(\"[\d]+\"])', dataWarning2)
                    if flag2 == 1:
                        MisraAdded = MisraAdded + '[' + misra[i].replace('\\','') + '] '
                        flag2 = 0
                dataWarning2 = dataWarning2.replace('],\n','],') 
                if flag == 1: 
                    print('\n[INFO] Position MISRA code %s: Added comment'%MisraAdded)
                    WorkingData = WorkingData.replace(dataWarning1,dataWarning2)
                    flag = 0 
                    MisraAdded = ''
                else:
                    print('\n[INFO] No empty warning to comment')
            else:
                print('[PROCESS] Ignore file '+ w_file + ': No warning')
        OutputFile = open(outputfolder + '\\temp\\' + w_file,'w')     
        OutputFile.write(WorkingData)
        OutputFile.close()
        
        
        
def add_Metric_comment(fin,fout,comment2,filename):
    flag = 0
    flag2 = 1
    chosen_index = 0
    j = 0
    a = 0
    count = 0
    print('[INFO] Scan HIS metric for',filename)
    for line in fin:
        #Get max values array
        if '[limit_max' in line:
            max_number = line.count('[')
            string_max = line.split('[', max_number)
            limit_max = string_max[max_number]
            commar_count = limit_max.count(',')                
            max_list = limit_max.split(',', commar_count)
        #Get min values array    
        if '[limit_min' in line:
            min_number = line.count('[')
            string_min = line.split('[', min_number)
            limit_min = string_min[min_number-1]
            commar_count = limit_min.count(',')                
            min_list = limit_min.split(',', commar_count)
        
        if '["FUNCTION"]' in line:
            number = line.count('["')
            string_3 = line.split('["', number)    
            for index, sub_string_3 in enumerate(string_3[2:number+1]):
                #print 'index' , index
                commar_count = sub_string_3.count(',"')
                char_3 = sub_string_3.split(',"', commar_count)
                for e in list:     
                    if len(char_3[e]) > 1:
                        #print char_3[e][0:len(char_3[e])-1]
                        #print char_3[e][1:len(char_3[e])-1]
                        a = float(char_3[e][0:len(char_3[e])-1])
                        b = float(max_list[e][1:len(max_list[e])-1])
                        c = float(min_list[e][1:len(min_list[e])-1])
                        if a > b or a < c:
                            char_3[e+1] = comment2[randint(1,row_count2)] +'\"'
                            flag = 1
                            chosen_index = index
                if flag == 1:
                    char_3_union = ',"'
                    flag = 0
                    char_3_union = char_3_union.join(char_3)
                    #print (chosen_index)
                    if 2+chosen_index <= number-1:
                        tail = line.index(string_3[2+chosen_index+1])
                        head = line.index(string_3[2+chosen_index])
                        flag2 = flag2 + 1
                        line = line[:head] + char_3_union + line[tail-2:]    
                    else:
                        head = line.index(string_3[2+chosen_index])
                        flag2 = flag2 + 1
                        line = line[:head] + char_3_union
        fout.write(line.encode())
    if flag2 != 1:
        print('\tFound ' + str(flag2) + ' blank HIS METRIC in',filename)
        print('\t\tAdded comment to blank cells')
    fout.close()

WorkingData = ''
row_count2 = 0
misra = {}
checkwarning = []
checkmetric = []
comment = {}
comment2 = {}
extend = {}
howtosolve = {}
list = [3, 19, 23, 43, 59, 69, 75, 81, 87, 99, 115, 127, 135]
 
print('\n\t\t\t\t\t _________________________________________ ')
print('\n\t\t\t\t\t|        QAC_Comment - Version 1.0        |')
print('\n\t\t\t\t\t|           Comment to QAC reports        |')
print('\n\t\t\t\t\t|              Author: Nguydu18           |')
print('\n\t\t\t\t\t ----------------------------------------- ')

# input #==================================================================================================================================================================
print("\nStarting...\n")
workspace = os.getcwd()
excelfile = os.path.normpath(workspace + os.sep + os.pardir)+ '\\QAC_Report_Review-Pro.xlsm'

if __name__ == "__main__":
    inputfolder = sys.argv[1]
inputfolder = inputfolder.replace(' ','')
outputfolder = os.path.normpath(inputfolder + os.sep + os.pardir)+ '\\QAC_Commented_File'
if os.path.exists(outputfolder):
    os.chmod(outputfolder, 0o777)
    shutil.rmtree(outputfolder)
os.makedirs(outputfolder)
os.makedirs(outputfolder + '\\temp')
print('Script working space:',workspace) 
print('QAC reports directory:',inputfolder) 
print('==========================================================================================')
if (path.exists(excelfile) == False):
    print('[ERROR] Excel file not found:',excelfile)
    input("Press Enter to close")
    exit()
if not any(fname.endswith('.c.html') for fname in os.listdir(inputfolder)):
    print('[ERROR] Not found any QAC report in',inputfolder) 
    input("Press Enter to close")
    exit()
# input #=========================================================================================

try:  
    #load the work book
    wb_obj = load_workbook(filename = excelfile)
    if 'warning' in wb_obj.sheetnames:
        wsheet = wb_obj['warning']
    else:
        print('[ERROR] Not found "warning" sheet in Excel file')
        input("Press Enter to close")
        exit() 
    row_count = wsheet.max_row
    # loop to get row and column values
    #GET WARNING COMMENT
    for i in range(2, row_count+1):
        misra[i]   = str(wsheet.cell(row=i, column=1).value)
        extend[i]  = wsheet.cell(row=i, column=2).value
        comment[i] = wsheet.cell(row=i, column=4).value	
        howtosolve[i] = wsheet.cell(row=i, column=3).value   
        if misra[i] != None:
            misra[i] = misra[i].replace('\\','\\\\')
            misra[i] = misra[i].replace('.','\.')     
        if howtosolve[i] != None:
            howtosolve[i] = howtosolve[i].replace('\\','\\\\')   
            howtosolve[i] = howtosolve[i].replace('"','')            
            howtosolve[i] = howtosolve[i].replace('(','\(')
            howtosolve[i] = howtosolve[i].replace(')','\)')            
    #GET METRIC COMMENT
    if 'metric' in wb_obj.sheetnames:
        wsheet = wb_obj['metric']
    else:
        print('[ERROR] Not found "metric" sheet in Excel file')
        input("Press Enter to close")
        exit()
    row_count2 = wsheet.max_row
    for i in range(1, row_count2+1):
        comment2[i] = wsheet.cell(row=i, column=2).value
    print('==========================================================================================')
    print('[PROCESS] Comment to QAC reports:')
    for i in range(2, row_count+1):
        if comment[i] != None:
            comment[i] = comment[i].replace('\\','\\\\')
            comment[i] = comment[i].replace('"','')
            comment[i] = comment[i].replace(')','\)')
            comment[i] = comment[i].replace('(','\(')
            howtosolve[i] = "" #Have comment > ignore fill warning to check
        else: 
            comment[i] = "" # comment[i] == none
    Review_Comment()
    #FixMetric:
    print('==========================================================================================\n')
    print('[PROCESS] Comment to HIS METRIC:')    
    for w_file in os.listdir(outputfolder + '\\temp'):
        fin = open(outputfolder + '\\temp\\' + w_file)
        fout = open(outputfolder + '\\' + w_file, 'ab')
        add_Metric_comment(fin,fout,comment2,w_file)
        fin.close()
    
    if os.path.exists(outputfolder + '\\temp'):
        shutil.rmtree(outputfolder + '\\temp')
    #===============================================================================================================================================================================================              
    print("\n\n\nTool run completed, pls check output at:",outputfolder)
except ValueError as outputError:
    print(outputError)  

os.system("pause")
















##reser
        #COMMENT METRIC
        # checkmetric  = re.search('(,\[\"[a-zA-Z_\-.]+\"+..??[0-9]......)(0\",\")([0-9,\".]+\])',WorkingData)
        # if checkmetric !=None:
            # countp = 0
            # metricData = ''
            # metricount = checkmetric.group(3)
            # # #print('REMETRIC =: ', ''.join(checkmetric.group(0)))
            # # #print('REMETRIC group3 =: ',metricount)
            # for x in range(0,len(metricount)):
                # if metricount[x] == ',':
                    # countp = countp + 1
                    # if countp == 93:
                        # break
            # # #print('vi tri can thay doi %s: voi gia tri la: '%x, metricount[x])
            # for y in range(0,x-1):
                # metricData = metricData + metricount[y]
            # metricData = metricData + comment2[0]
            # for y in range(x+3,len(metricount)):
                # metricData = metricData + metricount[y]
            # # # #print('gia tri bien metri sau khi gan', metricData)
            # # # #print('checkmetric.group(1): ', checkmetric.group(1))
            # metricData = checkmetric.group(1) + checkmetric.group(2) + comment2[0] + metricData
            # # # #print('gia tri bien moi tao', metricData)
            # while checkmetric != None:
                # concatresult = ''.join(checkmetric.group(0))                                             
                # WorkingData = WorkingData.replace(concatresult, metricData)
                # checkmetric = re.search('(,\[\"[a-zA-Z_\-.]+\"+..??[0-9]......)(0\",\")([0-9,\".]+\])',WorkingData)
            # print('\n[INFO] Metric comments was added to ',w_file)
